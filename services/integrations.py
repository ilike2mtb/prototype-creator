"""Core Figma API and SharePoint MS Graph functions.

Called by:
  - routers/figma.py and routers/sharepoint.py  → serves the Custom GPT HTTP endpoints
  - services/anthropic_service.py               → used by the LLM prototype pipeline

All Figma/SharePoint calls go directly to the upstream APIs (no HTTP self-calls).
"""

import io
from typing import Optional
from urllib.parse import quote

import httpx
from config import settings

FIGMA_API_BASE = "https://api.figma.com/v1"
GRAPH_API_BASE = "https://graph.microsoft.com/v1.0"

# Figma renders images on demand — give them extra time.
FIGMA_IMAGE_TIMEOUT = 60.0


def _timeout(read: Optional[float] = None) -> httpx.Timeout:
    return httpx.Timeout(read or settings.request_timeout_seconds, connect=30.0)


# ── Node / response trimming helpers ─────────────────────────────────────────


def _slim_figma_node(node: dict) -> dict:
    bounds = node.get("absoluteBoundingBox") if isinstance(node, dict) else None
    children = node.get("children")
    slim_children = []
    if isinstance(children, list):
        slim_children = [
            _slim_figma_node(child) for child in children if isinstance(child, dict)
        ]
    return {
        "id": node.get("id"),
        "name": node.get("name"),
        "type": node.get("type"),
        "visible": node.get("visible"),
        "characters": node.get("characters"),
        "layoutMode": node.get("layoutMode"),
        "itemSpacing": node.get("itemSpacing"),
        "paddingTop": node.get("paddingTop"),
        "paddingBottom": node.get("paddingBottom"),
        "paddingLeft": node.get("paddingLeft"),
        "paddingRight": node.get("paddingRight"),
        "componentId": node.get("componentId"),
        "x": bounds.get("x") if isinstance(bounds, dict) else None,
        "y": bounds.get("y") if isinstance(bounds, dict) else None,
        "width": bounds.get("width") if isinstance(bounds, dict) else None,
        "height": bounds.get("height") if isinstance(bounds, dict) else None,
        "children": slim_children,
    }


def _is_figma_error_response(data: dict) -> bool:
    return bool(data.get("error")) or "figma_status" in data


def _is_figma_variables_scope_error(data: dict) -> bool:
    if data.get("figma_status") != 403:
        return False
    body = str(data.get("figma_body") or "")
    return "file_variables:read" in body


def _trim_figma_file_response(data: dict) -> dict:
    if _is_figma_error_response(data):
        return data
    document = data.get("document")
    return {
        "name": data.get("name"),
        "lastModified": data.get("lastModified"),
        "version": data.get("version"),
        "document": _slim_figma_node(document) if isinstance(document, dict) else None,
    }


def _trim_figma_nodes_response(data: dict) -> dict:
    if _is_figma_error_response(data):
        return data
    nodes = data.get("nodes")
    trimmed_nodes: dict = {}
    if isinstance(nodes, dict):
        for node_id, node_data in nodes.items():
            if isinstance(node_data, dict) and isinstance(node_data.get("document"), dict):
                trimmed_nodes[node_id] = _slim_figma_node(node_data["document"])
            else:
                trimmed_nodes[node_id] = None
    return {"name": data.get("name"), "nodes": trimmed_nodes}


def _trim_figma_images_response(data: dict) -> dict:
    if _is_figma_error_response(data):
        return data
    trimmed = {"images": data.get("images", {})}
    if "err" in data:
        trimmed["err"] = data["err"]
    return trimmed


def _collect_page_names(document: dict) -> list[str]:
    children = document.get("children")
    if not isinstance(children, list):
        return []
    return [
        child.get("name") or child.get("id") or "Unnamed"
        for child in children
        if isinstance(child, dict)
    ]


def _collect_top_level_frame_names(document: dict) -> list[str]:
    children = document.get("children")
    if not isinstance(children, list):
        return []

    frame_names: list[str] = []
    for page in children:
        if not isinstance(page, dict):
            continue
        page_children = page.get("children")
        if not isinstance(page_children, list):
            continue
        for node in page_children:
            if isinstance(node, dict) and node.get("type") == "FRAME":
                frame_names.append(node.get("name") or node.get("id") or "Unnamed")
    return frame_names


# ── Frame metadata helpers ────────────────────────────────────────────────────


def _parse_figma_ids(ids: Optional[str]) -> list[str]:
    if not ids:
        return []
    parsed: list[str] = []
    for raw_part in ids.split(","):
        part = raw_part.strip()
        if part:
            parsed.append(part.replace("-", ":"))
    return parsed


def _frame_metadata_from_node(node: dict, path: list[str], parent_id: Optional[str]) -> dict:
    bounds = node.get("absoluteBoundingBox") if isinstance(node, dict) else None
    return {
        "id": node.get("id"),
        "name": node.get("name"),
        "type": node.get("type"),
        "path": " / ".join(path),
        "parentId": parent_id,
        "childCount": (
            len(node.get("children", [])) if isinstance(node.get("children"), list) else 0
        ),
        "layoutMode": node.get("layoutMode"),
        "itemSpacing": node.get("itemSpacing"),
        "paddingTop": node.get("paddingTop"),
        "paddingBottom": node.get("paddingBottom"),
        "paddingLeft": node.get("paddingLeft"),
        "paddingRight": node.get("paddingRight"),
        "x": bounds.get("x") if isinstance(bounds, dict) else None,
        "y": bounds.get("y") if isinstance(bounds, dict) else None,
        "width": bounds.get("width") if isinstance(bounds, dict) else None,
        "height": bounds.get("height") if isinstance(bounds, dict) else None,
    }


def _collect_frame_metadata(
    node: dict,
    path: Optional[list[str]] = None,
    parent_id: Optional[str] = None,
) -> list[dict]:
    if not isinstance(node, dict):
        return []
    current_path = (path or []) + [node.get("name") or node.get("id") or "Unnamed"]
    frames: list[dict] = []
    if node.get("type") == "FRAME":
        frames.append(_frame_metadata_from_node(node, current_path, parent_id))
    children = node.get("children")
    if isinstance(children, list):
        node_id = node.get("id") if isinstance(node.get("id"), str) else parent_id
        for child in children:
            if isinstance(child, dict):
                frames.extend(_collect_frame_metadata(child, current_path, node_id))
    return frames


def _search_figma_nodes(
    node: dict,
    query: str,
    node_type: Optional[str] = None,
    page_name: Optional[str] = None,
    path: Optional[list[str]] = None,
    parent_id: Optional[str] = None,
    current_page_name: Optional[str] = None,
) -> list[dict]:
    if not isinstance(node, dict):
        return []

    current_name = node.get("name") or node.get("id") or "Unnamed"
    current_path = (path or []) + [current_name]
    resolved_page_name = current_page_name
    if node.get("type") == "CANVAS":
        resolved_page_name = node.get("name") or current_page_name

    normalized_query = query.strip().lower()
    normalized_node_type = (node_type or "").strip().upper()
    normalized_page_name = (page_name or "").strip().lower()

    haystacks = [
        str(node.get("name") or "").lower(),
        str(node.get("characters") or "").lower(),
        str(node.get("id") or "").lower(),
    ]
    query_matches = any(normalized_query in haystack for haystack in haystacks if haystack)
    type_matches = not normalized_node_type or str(node.get("type") or "").upper() == normalized_node_type
    page_matches = not normalized_page_name or str(resolved_page_name or "").lower() == normalized_page_name

    matches: list[dict] = []
    bounds = node.get("absoluteBoundingBox")
    if query_matches and type_matches and page_matches:
        matches.append(
            {
                "id": node.get("id"),
                "name": node.get("name"),
                "type": node.get("type"),
                "pageName": resolved_page_name,
                "path": " / ".join(current_path),
                "parentId": parent_id,
                "textPreview": (node.get("characters") or "")[:200] or None,
                "componentId": node.get("componentId"),
                "x": bounds.get("x") if isinstance(bounds, dict) else None,
                "y": bounds.get("y") if isinstance(bounds, dict) else None,
                "width": bounds.get("width") if isinstance(bounds, dict) else None,
                "height": bounds.get("height") if isinstance(bounds, dict) else None,
            }
        )

    children = node.get("children")
    if isinstance(children, list):
        node_id = node.get("id") if isinstance(node.get("id"), str) else parent_id
        for child in children:
            if isinstance(child, dict):
                matches.extend(
                    _search_figma_nodes(
                        child,
                        query=query,
                        node_type=node_type,
                        page_name=page_name,
                        path=current_path,
                        parent_id=node_id,
                        current_page_name=resolved_page_name,
                    )
                )
    return matches


def _extract_frame_ids_from_frames(frames: list[dict]) -> list[str]:
    frame_ids: list[str] = []
    seen: set[str] = set()
    for frame in frames:
        frame_id = frame.get("id")
        if isinstance(frame_id, str) and frame_id and frame_id not in seen:
            seen.add(frame_id)
            frame_ids.append(frame_id)
    return frame_ids


def _chunk_list(values: list[str], chunk_size: int) -> list[list[str]]:
    if chunk_size < 1:
        chunk_size = 1
    return [values[i: i + chunk_size] for i in range(0, len(values), chunk_size)]


# ── Default resolution helpers ────────────────────────────────────────────────


def _resolve_file_key(file_key: Optional[str] = None) -> Optional[str]:
    """Return the first non-empty file key: query param → FIGMA_FILE_KEY_2 → FIGMA_FILE_KEY."""
    return (file_key or "").strip() or settings.figma_file_key_2 or settings.figma_file_key or None


def _resolve_node_ids(ids: Optional[str] = None) -> Optional[str]:
    """Return node IDs, normalising dash→colon format."""
    raw = (ids or "").strip() or settings.figma_node_ids
    if not raw:
        return None
    return ",".join(part.strip().replace("-", ":") for part in raw.split(",") if part.strip())


def _resolve_image_ids(ids: Optional[str] = None) -> Optional[str]:
    raw = (ids or "").strip() or settings.figma_image_ids
    if not raw:
        return None
    return ",".join(part.strip().replace("-", ":") for part in raw.split(",") if part.strip())


def _resolve_depth(depth: Optional[int] = None) -> int:
    return depth if depth is not None else settings.figma_depth


def _resolve_format(format_val: Optional[str] = None) -> str:
    fmt = (format_val or settings.figma_image_format or "png").strip()
    if fmt not in {"jpg", "png", "svg", "pdf"}:
        fmt = "png"
    return fmt


def _resolve_scale(scale_val: Optional[float] = None) -> float:
    return scale_val if scale_val is not None else settings.figma_image_scale


# ── Figma API HTTP helper ─────────────────────────────────────────────────────


async def _figma_get(path: str, params: Optional[dict] = None,
                     read_timeout: Optional[float] = None) -> dict:
    if not settings.figma_token:
        return {"error": "FIGMA_TOKEN is not configured"}
    headers = {"X-Figma-Token": settings.figma_token}
    if settings.debug:
        print("FIGMA GET:", path, params)
    async with httpx.AsyncClient(timeout=_timeout(read_timeout)) as client:
        response = await client.get(
            f"{FIGMA_API_BASE}{path}", headers=headers, params=params or {}
        )
        if response.status_code >= 400:
            detail = {"figma_status": response.status_code, "figma_body": response.text}
            if settings.debug:
                print("FIGMA ERROR:", detail)
            return detail
        return response.json()


# ── Public Figma functions ────────────────────────────────────────────────────


async def get_file(file_key: Optional[str] = None) -> dict:
    fk = _resolve_file_key(file_key)
    data = await _figma_get(f"/files/{fk}")
    return _trim_figma_file_response(data)


async def get_file_summary(file_key: Optional[str] = None) -> dict:
    fk = _resolve_file_key(file_key)

    file_data = await _figma_get(f"/files/{fk}")
    if _is_figma_error_response(file_data):
        return file_data

    document = file_data.get("document")
    page_names = _collect_page_names(document) if isinstance(document, dict) else []
    top_level_frame_names = (
        _collect_top_level_frame_names(document) if isinstance(document, dict) else []
    )
    all_frames = _collect_frame_metadata(document) if isinstance(document, dict) else []

    components_data = await _figma_get(f"/files/{fk}/components")
    if _is_figma_error_response(components_data):
        return components_data

    styles_data = await _figma_get(f"/files/{fk}/styles")
    if _is_figma_error_response(styles_data):
        return styles_data

    variables_data = await _figma_get(f"/files/{fk}/variables/local")
    variables_unavailable = False
    variables_error = None
    variable_count = None
    if _is_figma_variables_scope_error(variables_data):
        variables_unavailable = True
        variables_error = "Figma token lacks file_variables:read scope."
    elif _is_figma_error_response(variables_data):
        return variables_data
    component_count = len(((components_data.get("meta") or {}).get("components") or []))
    style_count = len(((styles_data.get("meta") or {}).get("styles") or []))
    if not variables_unavailable:
        variable_count = len(((variables_data.get("meta") or {}).get("variables") or {}))

    return {
        "name": file_data.get("name"),
        "lastModified": file_data.get("lastModified"),
        "version": file_data.get("version"),
        "pageNames": page_names,
        "topLevelFrameNames": top_level_frame_names,
        "totalPages": len(page_names),
        "totalTopLevelFrames": len(top_level_frame_names),
        "totalFrames": len(all_frames),
        "componentCount": component_count,
        "styleCount": style_count,
        "variableCount": variable_count,
        "variablesUnavailable": variables_unavailable,
        "variablesError": variables_error,
    }


async def get_nodes(
    file_key: Optional[str] = None,
    ids: Optional[str] = None,
    depth: Optional[int] = None,
) -> dict:
    fk = _resolve_file_key(file_key)
    nids = _resolve_node_ids(ids)
    d = _resolve_depth(depth)
    data = await _figma_get(f"/files/{fk}/nodes", {"ids": nids, "depth": d})
    return _trim_figma_nodes_response(data)


async def get_frames(
    file_key: Optional[str] = None,
    ids: Optional[str] = None,
    depth: Optional[int] = None,
) -> dict:
    fk = _resolve_file_key(file_key)
    nids = _resolve_node_ids(ids)
    d = _resolve_depth(depth)
    data = await _figma_get(f"/files/{fk}/nodes", {"ids": nids, "depth": d})
    if _is_figma_error_response(data):
        return data

    frames: list[dict] = []
    nodes = data.get("nodes")
    if isinstance(nodes, dict):
        for node_data in nodes.values():
            if isinstance(node_data, dict) and isinstance(node_data.get("document"), dict):
                frames.extend(_collect_frame_metadata(node_data["document"]))

    return {
        "name": data.get("name"),
        "requestedIds": nids,
        "depth": d,
        "totalFrames": len(frames),
        "frames": frames,
    }


async def search_file(
    query: str,
    file_key: Optional[str] = None,
    node_type: Optional[str] = None,
    page_name: Optional[str] = None,
    limit: int = 25,
) -> dict:
    fk = _resolve_file_key(file_key)
    file_data = await _figma_get(f"/files/{fk}")
    if _is_figma_error_response(file_data):
        return file_data

    document = file_data.get("document")
    matches = _search_figma_nodes(
        document,
        query=query,
        node_type=node_type,
        page_name=page_name,
    ) if isinstance(document, dict) else []

    return {
        "name": file_data.get("name"),
        "query": query,
        "nodeType": node_type,
        "pageName": page_name,
        "totalMatches": len(matches),
        "limit": limit,
        "matches": matches[:limit],
    }


async def get_file_images(
    file_key: Optional[str] = None,
    ids: Optional[str] = None,
    format: Optional[str] = None,
    scale: Optional[float] = None,
) -> dict:
    """Get images for specific node IDs. Used by /figma/file/images endpoint."""
    fk = _resolve_file_key(file_key)
    nids = _resolve_image_ids(ids)
    fmt = _resolve_format(format)
    scl = _resolve_scale(scale)
    data = await _figma_get(f"/images/{fk}", {"ids": nids, "format": fmt, "scale": scl},
                            read_timeout=FIGMA_IMAGE_TIMEOUT)
    return _trim_figma_images_response(data)


async def export_images(
    file_key: Optional[str] = None,
    ids: Optional[str] = None,
    format: Optional[str] = None,
    scale: Optional[float] = None,
) -> dict:
    """Export rendered images for frame IDs.

    If ids is None/empty, discovers all frames in the file first.
    Used by both the LLM pipeline (with explicit ids) and the /figma/file/frames/images endpoint.
    """
    fk = _resolve_file_key(file_key)
    fmt = _resolve_format(format)
    scl = _resolve_scale(scale)
    frame_ids = _parse_figma_ids(ids) if ids else []

    if not frame_ids:
        # Discover all frames in the file
        file_data = await _figma_get(f"/files/{fk}")
        document = file_data.get("document")
        if isinstance(document, dict):
            frames = _collect_frame_metadata(document)
            frame_ids = _extract_frame_ids_from_frames(frames)

    if not frame_ids:
        return {"totalFrames": 0, "frames": [], "images": {}}

    batch_size = settings.figma_image_batch_size
    batches = _chunk_list(frame_ids, batch_size)

    merged_images: dict[str, str] = {}
    merged_errors: list[str] = []
    for batch in batches:
        images_data = await _figma_get(
            f"/images/{fk}",
            {"ids": ",".join(batch), "format": fmt, "scale": scl},
            read_timeout=FIGMA_IMAGE_TIMEOUT,
        )
        trimmed = _trim_figma_images_response(images_data)
        batch_images = trimmed.get("images", {})
        if isinstance(batch_images, dict):
            merged_images.update(batch_images)
        if trimmed.get("err"):
            merged_errors.append(str(trimmed["err"]))

    result: dict = {
        "totalFrames": len(frame_ids),
        "frameIds": frame_ids,
        "images": merged_images,
        "batchSize": batch_size,
        "batchCount": len(batches),
    }
    if merged_errors:
        result["err"] = " | ".join(merged_errors)
    return result


# ── Design context helpers ────────────────────────────────────────────────────


def _rgba_to_hex(color: dict) -> str:
    """Convert a Figma RGBA color dict {r,g,b,a} (0-1 floats) to a CSS hex string."""
    r = int(color.get("r", 0) * 255)
    g = int(color.get("g", 0) * 255)
    b = int(color.get("b", 0) * 255)
    return f"#{r:02x}{g:02x}{b:02x}"


def _slim_variables_response(data: dict) -> dict:
    meta        = data.get("meta") or {}
    collections = meta.get("variableCollections") or {}
    variables   = meta.get("variables") or {}

    col_names = {cid: (col.get("name") or cid) for cid, col in collections.items()}

    slim: list[dict] = []
    for var in variables.values():
        resolved_type = var.get("resolvedType")
        entry: dict = {
            "name":       var.get("name"),
            "type":       resolved_type,
            "collection": col_names.get(var.get("variableCollectionId") or "", ""),
        }
        values_by_mode = var.get("valuesByMode") or {}
        if resolved_type == "COLOR":
            col_id       = var.get("variableCollectionId") or ""
            default_mode = (collections.get(col_id) or {}).get("defaultModeId")
            color        = (
                values_by_mode.get(default_mode)
                if default_mode
                else next(iter(values_by_mode.values()), None)
            )
            if isinstance(color, dict) and "r" in color:
                entry["hex"] = _rgba_to_hex(color)
        else:
            first_value = next(iter(values_by_mode.values()), None)
            if first_value is not None:
                entry["value"] = str(first_value)
        slim.append(entry)

    return {
        "variableCount": len(slim),
        "collections":   list(col_names.values()),
        "variables":     slim,
    }


def _slim_components_response(data: dict) -> dict:
    components = (data.get("meta") or {}).get("components") or []
    slim = [
        {
            "name":           c.get("name"),
            "description":    c.get("description") or "",
            "nodeId":         c.get("nodeId"),
            "componentSetId": c.get("componentSetId"),
        }
        for c in components
    ]
    return {"componentCount": len(slim), "components": slim}


def _slim_styles_response(data: dict) -> dict:
    styles = (data.get("meta") or {}).get("styles") or []
    slim = [
        {
            "name":        s.get("name"),
            "type":        s.get("styleType"),   # FILL | TEXT | EFFECT | GRID
            "description": s.get("description") or "",
            "nodeId":      s.get("nodeId"),
        }
        for s in styles
    ]
    return {"styleCount": len(slim), "styles": slim}


# ── Public design-context Figma functions ─────────────────────────────────────


async def get_variables(file_key: Optional[str] = None) -> dict:
    """Get local variables (design tokens — colors, spacing, type scales) from a Figma file."""
    fk   = _resolve_file_key(file_key)
    data = await _figma_get(f"/files/{fk}/variables/local")
    if _is_figma_error_response(data):
        return data
    return _slim_variables_response(data)


async def get_components(file_key: Optional[str] = None) -> dict:
    """Get the component library (names, descriptions, node IDs) from a Figma file."""
    fk   = _resolve_file_key(file_key)
    data = await _figma_get(f"/files/{fk}/components")
    if _is_figma_error_response(data):
        return data
    return _slim_components_response(data)


async def get_styles(file_key: Optional[str] = None) -> dict:
    """Get published styles (color, text, effect, grid) from a Figma file."""
    fk   = _resolve_file_key(file_key)
    data = await _figma_get(f"/files/{fk}/styles")
    if _is_figma_error_response(data):
        return data
    return _slim_styles_response(data)


# ── SharePoint / MS Graph functions ──────────────────────────────────────────


async def _graph_access_token(client: httpx.AsyncClient) -> str:
    token_url = (
        f"https://login.microsoftonline.com/{settings.ms_tenant_id}/oauth2/v2.0/token"
    )
    payload = {
        "client_id": settings.ms_client_id,
        "client_secret": settings.ms_client_secret,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }
    response = await client.post(token_url, data=payload)
    if response.status_code >= 400:
        raise Exception(
            f"Graph auth failed: {response.status_code} {response.text[:200]}"
        )
    token = response.json().get("access_token")
    if not token:
        raise Exception("No Graph access token returned")
    return token


async def _graph_download_binary(
    client: httpx.AsyncClient, token: str, path: str
) -> bytes:
    headers = {"Authorization": f"Bearer {token}"}
    response = await client.get(
        f"{GRAPH_API_BASE}{path}", headers=headers, follow_redirects=True
    )
    if response.status_code >= 400:
        raise Exception(
            f"Graph download failed: {response.status_code} {response.text[:200]}"
        )
    return response.content


def _normalize_cell(value: object) -> str:
    return "" if value is None else str(value)


def _values_to_rows(
    values: list[list[object]],
) -> tuple[list[str], list[dict[str, str]]]:
    if not values:
        return [], []
    raw_headers = values[0]
    headers: list[str] = []
    header_counts: dict[str, int] = {}
    for index, raw_header in enumerate(raw_headers):
        header = _normalize_cell(raw_header).strip() or f"column_{index + 1}"
        if header in header_counts:
            header_counts[header] += 1
            header = f"{header}_{header_counts[header]}"
        else:
            header_counts[header] = 1
        headers.append(header)
    rows: list[dict[str, str]] = []
    for raw_row in values[1:]:
        row_values = [_normalize_cell(item) for item in raw_row]
        if not any(cell.strip() for cell in row_values):
            continue
        row: dict[str, str] = {}
        for idx, header in enumerate(headers):
            row[header] = row_values[idx] if idx < len(row_values) else ""
        rows.append(row)
    return headers, rows


async def _fetch_sharepoint_sheet(file_path_raw: str) -> dict:
    from openpyxl import load_workbook

    path = quote(file_path_raw.lstrip("/"), safe="/")
    file_path = f"/drives/{settings.sp_drive_id}/root:/{path}:/content"

    async with httpx.AsyncClient(timeout=_timeout()) as client:
        token = await _graph_access_token(client)
        file_bytes = await _graph_download_binary(client, token, file_path)

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if settings.sp_worksheet_name:
        if settings.sp_worksheet_name not in workbook.sheetnames:
            raise Exception(f"Worksheet '{settings.sp_worksheet_name}' not found")
        worksheet = workbook[settings.sp_worksheet_name]
    else:
        worksheet = workbook.active

    values = list(worksheet.values)
    if not values:
        raise Exception("Spreadsheet contains no data")

    headers, rows = _values_to_rows(values)
    return {
        "file_name": file_path_raw.split("/")[-1],
        "file_web_url": None,
        "worksheet": worksheet.title,
        "headers": headers,
        "total_rows": len(rows),
        "rows": rows,
    }


async def get_architecture_plan() -> dict:
    """Fetch the DCI architecture plan from SharePoint (example file)."""
    return await _fetch_sharepoint_sheet(settings.sp_file_path_example)


async def get_architecture_template() -> dict:
    """Fetch the architecture plan template from SharePoint."""
    return await _fetch_sharepoint_sheet(settings.sp_file_path)
