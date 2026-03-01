import os
import io
from typing import Annotated, Optional
from urllib.parse import quote

import httpx
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.openapi.utils import get_openapi
from fastapi.responses import JSONResponse, Response
from fastapi.security import APIKeyHeader
from openpyxl import load_workbook
from pydantic import BaseModel

load_dotenv()

DEBUG = os.getenv("DEBUG", "False").lower() == "true"

FIGMA_API_BASE = "https://api.figma.com/v1"
GRAPH_API_BASE = "https://graph.microsoft.com/v1.0"
GRAPH_BASE_URL = GRAPH_API_BASE

FIGMA_TOKEN = os.getenv("FIGMA_TOKEN", "").strip()
FIGMA_FILE_KEY = os.getenv("FIGMA_FILE_KEY_2", "").strip()
FIGMA_NODE_IDS = os.getenv("FIGMA_NODE_IDS", "").strip()
FIGMA_IMAGE_IDS = os.getenv("FIGMA_IMAGE_IDS", "").strip()
FIGMA_DEPTH = int(os.getenv("FIGMA_DEPTH", "2"))
SERVICE_KEY = os.getenv("SERVICE_KEY", "").strip()
REQUEST_TIMEOUT_SECONDS = float(os.getenv("REQUEST_TIMEOUT_SECONDS", "20"))
PUBLIC_BASE_URL = (
    os.getenv("PUBLIC_BASE_URL")
    or os.getenv("RENDER_EXTERNAL_URL")
    or ""
).strip()

MS_TENANT_ID = os.getenv("MS_TENANT_ID", "").strip()
MS_CLIENT_ID = os.getenv("MS_CLIENT_ID", "").strip()
MS_CLIENT_SECRET = os.getenv("MS_CLIENT_SECRET", "").strip()
SP_SITE_ID = os.getenv("SP_SITE_ID", "").strip()
SP_DRIVE_ID = os.getenv("SP_DRIVE_ID", "").strip()
SP_FILE_PATH = os.getenv("SP_FILE_PATH", "").strip()
SP_WORKSHEET_NAME = os.getenv("SP_WORKSHEET_NAME", "").strip()

app = FastAPI(
    title="Figma Proxy API",
    version="1.0.0",
    description="Small authenticated API for Custom GPT actions to query Figma.",
)

service_key_header = APIKeyHeader(name="X-Service-Key", auto_error=False)


class HealthResponse(BaseModel):
    ok: bool


class RootResponse(BaseModel):
    name: str
    ok: bool
    openapi: str
    docs: str


class SharePointSheetResponse(BaseModel):
    file_name: str
    file_web_url: Optional[str] = None
    worksheet: str
    headers: list[str]
    total_rows: int
    rows: list[dict[str, str]]


def custom_openapi() -> dict:
    if app.openapi_schema:
        return app.openapi_schema
    openapi_schema = get_openapi(
        title=app.title,
        version=app.version,
        description=app.description,
        routes=app.routes,
    )
    openapi_schema["openapi"] = "3.1.0"
    if PUBLIC_BASE_URL:
        openapi_schema["servers"] = [{"url": PUBLIC_BASE_URL}]
    app.openapi_schema = openapi_schema
    return app.openapi_schema


app.openapi = custom_openapi


def _require_service_key(x_service_key: Optional[str]) -> None:
    if not SERVICE_KEY:
        return
    if not x_service_key or x_service_key != SERVICE_KEY:
        raise HTTPException(status_code=401, detail="Invalid service key")


def require_service_key(
    x_service_key: Annotated[Optional[str], Depends(service_key_header)],
) -> None:
    _require_service_key(x_service_key)


def _require_figma_token() -> None:
    if not FIGMA_TOKEN:
        raise HTTPException(status_code=500, detail="FIGMA_TOKEN is not configured")


def _get_figma_file_key() -> str:
    if not FIGMA_FILE_KEY:
        raise HTTPException(status_code=500, detail="FIGMA_FILE_KEY is not configured")
    return FIGMA_FILE_KEY


def _get_figma_node_ids(ids_override: Optional[str] = None) -> str:
    if ids_override and ids_override.strip():
        return ids_override.strip()
    if not FIGMA_NODE_IDS:
        raise HTTPException(
            status_code=500,
            detail="FIGMA_NODE_IDS is not configured and no ids query parameter was provided",
        )
    return FIGMA_NODE_IDS


def _get_figma_image_ids(ids_override: Optional[str] = None) -> str:
    if ids_override and ids_override.strip():
        return ids_override.strip()
    if not FIGMA_IMAGE_IDS:
        raise HTTPException(
            status_code=500,
            detail="FIGMA_IMAGE_IDS is not configured and no ids query parameter was provided",
        )
    raw = FIGMA_IMAGE_IDS
    # Normalize dash format (1-88) → colon format (1:88)
    normalized = ",".join(part.replace("-", ":") for part in raw.split(","))

    return normalized

def _get_figma_depth(depth_override: Optional[int] = None) -> int:
    if depth_override is not None:
        return depth_override
    return FIGMA_DEPTH


def _slim_figma_node(node: dict) -> dict:
    children = node.get("children")
    slim_children = []
    if isinstance(children, list):
        slim_children = [
            _slim_figma_node(child) for child in children if isinstance(child, dict)
        ]

    return {
        "id": node.get("id"),
        "name": node.get("name"),
        "type": node.get("type"),
        "layoutMode": node.get("layoutMode"),
        "itemSpacing": node.get("itemSpacing"),
        "paddingTop": node.get("paddingTop"),
        "paddingBottom": node.get("paddingBottom"),
        "paddingLeft": node.get("paddingLeft"),
        "paddingRight": node.get("paddingRight"),
        "children": slim_children,
    }


def _trim_figma_file_response(data: dict) -> dict:
    document = data.get("document")
    return {
        "name": data.get("name"),
        "lastModified": data.get("lastModified"),
        "version": data.get("version"),
        "document": _slim_figma_node(document) if isinstance(document, dict) else None,
    }


def _trim_figma_nodes_response(data: dict) -> dict:
    nodes = data.get("nodes")
    trimmed_nodes: dict = {}

    if isinstance(nodes, dict):
        for node_id, node_data in nodes.items():
            if isinstance(node_data, dict) and isinstance(node_data.get("document"), dict):
                trimmed_nodes[node_id] = _slim_figma_node(node_data["document"])
            else:
                trimmed_nodes[node_id] = node_data

    return {"name": data.get("name"), "nodes": trimmed_nodes}


def _trim_figma_images_response(data: dict) -> dict:
    trimmed = {"images": data.get("images", {})}
    if "err" in data:
        trimmed["err"] = data["err"]
    return trimmed


def _require_sharepoint_config() -> None:
    missing = []
    if not MS_TENANT_ID:
        missing.append("MS_TENANT_ID")
    if not MS_CLIENT_ID:
        missing.append("MS_CLIENT_ID")
    if not MS_CLIENT_SECRET:
        missing.append("MS_CLIENT_SECRET")
    if not SP_DRIVE_ID:
        missing.append("SP_DRIVE_ID")
    if not SP_FILE_PATH:
        missing.append("SP_FILE_PATH")
    if missing:
        raise HTTPException(
            status_code=500,
            detail={"message": "SharePoint configuration is missing", "missing": missing},
        )


async def _figma_get(path: str, params: Optional[dict] = None) -> dict:
    _require_figma_token()
    headers = {"X-Figma-Token": FIGMA_TOKEN}
    url = f"{FIGMA_API_BASE}{path}"
    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT_SECONDS) as client:
        response = await client.get(url, headers=headers, params=params)
        if response.status_code >= 400:
            detail = {
                "figma_status": response.status_code,
                "figma_body": response.text,
            }
            if DEBUG:
                print("FIGMA ERROR:", detail)
            raise HTTPException(status_code=502, detail=detail)
        return response.json()


async def _graph_access_token(client: httpx.AsyncClient) -> str:
    token_url = f"https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/token"
    payload = {
        "client_id": MS_CLIENT_ID,
        "client_secret": MS_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials",
    }
    response = await client.post(token_url, data=payload)
    if response.status_code >= 400:
        raise HTTPException(
            status_code=502,
            detail={"graph_status": response.status_code, "graph_body": response.text},
        )
    token = response.json().get("access_token")
    if not token:
        raise HTTPException(status_code=502, detail="No Graph access token returned")
    return token


async def _graph_get_json(
    client: httpx.AsyncClient,
    token: str,
    path: str,
    params: Optional[dict] = None,
) -> dict:
    headers = {"Authorization": f"Bearer {token}"}

    response = await client.get(
        f"{GRAPH_API_BASE}{path}",
        headers=headers,
        params=params,
        follow_redirects=True,
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=502,
            detail={
                "graph_status": response.status_code,
                "graph_body": response.text,
            },
        )

    return response.json()
async def _graph_download_binary(
    client: httpx.AsyncClient,
    token: str,
    path: str,
) -> bytes:
    headers = {"Authorization": f"Bearer {token}"}

    response = await client.get(
        f"{GRAPH_API_BASE}{path}",
        headers=headers,
        follow_redirects=True,  # REQUIRED for /content
    )

    if response.status_code >= 400:
        raise HTTPException(
            status_code=502,
            detail={
                "graph_status": response.status_code,
                "graph_body": response.text,
            },
        )

    return response.content

def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _values_to_rows(values: list[list[object]]) -> tuple[list[str], list[dict[str, str]]]:
    if not values:
        return [], []

    raw_headers = values[0]
    headers: list[str] = []
    header_counts: dict[str, int] = {}
    for index, raw_header in enumerate(raw_headers):
        header = _normalize_cell(raw_header).strip() or f"column_{index + 1}"
        if header in header_counts:
            header_counts[header] += 1
            header = f"{header}_{header_counts[header]}"
        else:
            header_counts[header] = 1
        headers.append(header)

    rows: list[dict[str, str]] = []
    for raw_row in values[1:]:
        row_values = [_normalize_cell(item) for item in raw_row]
        if not any(cell.strip() for cell in row_values):
            continue
        row: dict[str, str] = {}
        for idx, header in enumerate(headers):
            row[header] = row_values[idx] if idx < len(row_values) else ""
        rows.append(row)

    return headers, rows


@app.get("/health", response_model=HealthResponse)
def health() -> HealthResponse:
    return {"ok": True}


@app.get("/", response_model=RootResponse)
def root() -> RootResponse:
    return {
        "name": app.title,
        "ok": True,
        "openapi": "/openapi.json",
        "docs": "/docs",
    }


@app.get("/figma/file")
async def get_file(
    _: None = Depends(require_service_key),
) -> JSONResponse:
    file_key = _get_figma_file_key()
    data = await _figma_get(f"/files/{file_key}")
    return JSONResponse(content=_trim_figma_file_response(data))


@app.get("/figma/file/nodes")
async def get_file_nodes(
    _: None = Depends(require_service_key),
    ids: Optional[str] = Query(
        default=None,
        description="Comma-separated node IDs. Optional if FIGMA_NODE_IDS is configured.",
    ),
    depth: Optional[int] = Query(default=None, ge=1, le=10),
) -> JSONResponse:
    file_key = _get_figma_file_key()
    params = {"ids": _get_figma_node_ids(ids)}
    params["depth"] = _get_figma_depth(depth)
    data = await _figma_get(f"/files/{file_key}/nodes", params=params)
    return JSONResponse(content=_trim_figma_nodes_response(data))


@app.api_route("/figma/file/images", methods=["GET", "POST"])
async def get_file_images(
    _: None = Depends(require_service_key),
    ids: Optional[str] = Query(
        default=None,
        description="Comma-separated node IDs. Optional if FIGMA_IMAGE_IDS is configured.",
    ),
    format: str = Query(default="png", pattern="^(jpg|png|svg|pdf)$"),
    scale: float = Query(default=1.0, ge=0.01, le=4.0),
) -> JSONResponse:
    file_key = _get_figma_file_key()
    params = {"ids": _get_figma_image_ids(ids), "format": format, "scale": scale}
    data = await _figma_get(f"/images/{file_key}", params=params)
    return JSONResponse(content=_trim_figma_images_response(data))

@app.get(
    "/sharepoint/dci-architecture-plan",
    response_model=SharePointSheetResponse,
)
async def get_dci_architecture_plan(
    _: None = Depends(require_service_key),
) -> SharePointSheetResponse:
    _require_sharepoint_config()

    # Build safe relative path
    path = quote(SP_FILE_PATH.lstrip("/"), safe="/")

    # Use drive-only resolution (no /sites/, no /workbook/)
    file_path = f"/drives/{SP_DRIVE_ID}/root:/{path}:/content"

    async with httpx.AsyncClient(timeout=REQUEST_TIMEOUT_SECONDS) as client:
        token = await _graph_access_token(client)

        # Download Excel file binary
        try:
            file_bytes = await _graph_download_binary(
                client,
                token,
                file_path,
            )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(
                status_code=502,
                detail=f"Failed to download SharePoint file: {str(e)}",
            )

    # Parse Excel locally (GPT-safe, no WAC)
    try:
        workbook = load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
            read_only=True,
        )

        if SP_WORKSHEET_NAME:
            if SP_WORKSHEET_NAME not in workbook.sheetnames:
                raise HTTPException(
                    status_code=400,
                    detail=f"Worksheet '{SP_WORKSHEET_NAME}' not found.",
                )
            worksheet = workbook[SP_WORKSHEET_NAME]
        else:
            worksheet = workbook.active

        values = list(worksheet.values)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=502,
            detail=f"Failed to parse Excel file: {str(e)}",
        )

    if not values:
        raise HTTPException(
            status_code=502,
            detail="Spreadsheet contains no data.",
        )

    headers, rows = _values_to_rows(values)

    return {
        "file_name": SP_FILE_PATH.split("/")[-1],
        "file_web_url": None,  # Optional — remove if not needed
        "worksheet": worksheet.title,
        "headers": headers,
        "total_rows": len(rows),
        "rows": rows,
    }
@app.middleware("http")
async def log_every_request(request: Request, call_next):
    print("INCOMING:", request.method, request.url.path)
    response = await call_next(request)
    print("STATUS:", response.status_code)
    return response
